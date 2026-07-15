# generate_load_report.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_load_report(metrics, logs):
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Summary Dashboard ─────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Load Test Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Styling
    brand_blue = "0ea5e9" # Tailwind sky-500
    brand_light_blue = "e0f2fe" # Tailwind sky-100
    success_green = "16a34a" # Tailwind green-600
    success_light = "dcfce7"
    white = "FFFFFF"
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color=white)
    title_font = Font(name="Segoe UI", size=16, bold=True, color="0f172a")
    section_font = Font(name="Segoe UI", size=13, bold=True, color="1e293b")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)
    
    header_fill = PatternFill(start_color=brand_blue, end_color=brand_blue, fill_type="solid")
    light_blue_fill = PatternFill(start_color=brand_light_blue, end_color=brand_light_blue, fill_type="solid")
    green_fill = PatternFill(start_color=success_light, end_color=success_light, fill_type="solid")
    
    thin_gray = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    # Title
    ws_summary["A1"] = "ClinLab Application Baseline Load Test Report"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:D1")
    ws_summary.row_dimensions[1].height = 36

    # Configurations Section
    ws_summary["A3"] = "Test Configuration"
    ws_summary["A3"].font = section_font
    
    config_headers = ["Parameters", "Value"]
    for col_idx, h in enumerate(config_headers, start=1):
        cell = ws_summary.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all
        
    configs = [
        ("Target URL", metrics["target_url"]),
        ("Backend API URL", metrics["backend_url"]),
        ("Virtual Users (VU)", metrics["virtual_users"]),
        ("Test Duration (sec)", metrics["duration"])
    ]
    
    for idx, (param, val) in enumerate(configs, start=5):
        c1 = ws_summary.cell(row=idx, column=1, value=param)
        c2 = ws_summary.cell(row=idx, column=2, value=val)
        c1.font = regular_font
        c2.font = bold_font
        c1.border = border_all
        c2.border = border_all
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws_summary.row_dimensions[idx].height = 20

    # Metrics Section
    ws_summary["A11"] = "Execution Performance Summary"
    ws_summary["A11"].font = section_font
    
    metrics_headers = ["Key Performance Indicators (KPI)", "Result"]
    for col_idx, h in enumerate(metrics_headers, start=1):
        cell = ws_summary.cell(row=12, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all
        
    kpi_list = [
        ("Total Requests Sent", metrics["total_requests"]),
        ("Requests Per Second (RPS)", f"{metrics['rps']} req/sec"),
        ("Successful Requests (2xx)", metrics["success_count"]),
        ("Failed Requests (0/5xx/4xx)", metrics["failed_count"]),
        ("Overall Pass Rate", metrics["pass_rate"]),
        ("Average Latency (ms)", f"{metrics['avg_time']} ms"),
        ("Minimum Latency (ms)", f"{metrics['min_time']} ms"),
        ("Maximum Latency (ms)", f"{metrics['max_time']} ms")
    ]
    
    for idx, (kpi, val) in enumerate(kpi_list, start=13):
        c1 = ws_summary.cell(row=idx, column=1, value=kpi)
        c2 = ws_summary.cell(row=idx, column=2, value=val)
        c1.font = regular_font
        c2.font = bold_font
        c1.border = border_all
        c2.border = border_all
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        ws_summary.row_dimensions[idx].height = 20
        if "Pass Rate" in kpi or "RPS" in kpi:
            c1.fill = green_fill
            c2.fill = green_fill
            
    # Auto-adjust column widths for summary sheet
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ── Sheet 2: Request Details Log ─────────────────────────────────────────
    ws_log = wb.create_sheet(title="Request Details Log")
    ws_log.views.sheetView[0].showGridLines = True
    
    log_headers = ["User ID", "Request #", "Endpoint", "HTTP Status", "Response Time (ms)", "Timestamp"]
    for col_idx, h in enumerate(log_headers, start=1):
        cell = ws_log.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
        
    ws_log.row_dimensions[1].height = 24
    
    status_pass_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    status_fail_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    
    for row_idx, log in enumerate(logs, start=2):
        ws_log.cell(row=row_idx, column=1, value=f"User #{log['user_id']}").alignment = Alignment(horizontal="center")
        ws_log.cell(row=row_idx, column=2, value=log["request_num"]).alignment = Alignment(horizontal="center")
        ws_log.cell(row=row_idx, column=3, value=log["endpoint"])
        
        status_cell = ws_log.cell(row=row_idx, column=4, value=log["status"])
        status_cell.alignment = Alignment(horizontal="center")
        if 200 <= log["status"] < 400:
            status_cell.fill = status_pass_fill
        else:
            status_cell.fill = status_fail_fill
            
        ws_log.cell(row=row_idx, column=5, value=log["duration"]).alignment = Alignment(horizontal="right")
        ws_log.cell(row=row_idx, column=6, value=log["timestamp"]).alignment = Alignment(horizontal="center")
        
        for col_idx in range(1, 7):
            c = ws_log.cell(row=row_idx, column=col_idx)
            c.font = regular_font
            c.border = border_all
            
        ws_log.row_dimensions[row_idx].height = 18
        
    # Auto column widths for details
    column_widths = {"A": 15, "B": 15, "C": 20, "D": 15, "E": 22, "F": 22}
    for col_letter, width in column_widths.items():
        ws_log.column_dimensions[col_letter].width = width

    # Save logic with locks handling
    try:
        wb.save("load_test.xlsx")
        print("[SUCCESS] Report generated: load_test.xlsx")
    except PermissionError:
        try:
            wb.save("load_test_report.xlsx")
            print("[WARNING] 'load_test.xlsx' is open. Saved as 'load_test_report.xlsx' instead.")
        except PermissionError:
            wb.save("clinlab_load_test.xlsx")
            print("[WARNING] Files are locked. Saved as 'clinlab_load_test.xlsx' instead.")

if __name__ == "__main__":
    # Test stub
    test_metrics = {
        "target_url": "https://clinlab-ai-assist.vercel.app",
        "backend_url": "https://pdd-backend-ztqc.onrender.com",
        "virtual_users": 100,
        "duration": 60,
        "total_requests": 8450,
        "success_count": 8450,
        "failed_count": 0,
        "pass_rate": "100.0%",
        "rps": 140.8,
        "avg_time": 210,
        "min_time": 45,
        "max_time": 1250
    }
    test_logs = [
        {"user_id": 1, "request_num": 1, "endpoint": "GET /", "status": 200, "duration": 210, "timestamp": "2026-07-14T14:44:00"}
    ]
    generate_load_report(test_metrics, test_logs)
